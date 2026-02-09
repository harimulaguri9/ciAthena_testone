from openpyxl import Workbook

class ExcelWriter:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.row = 1
        # Write header
        self.sheet.cell(row=1, column=1, value="Question")
        self.sheet.cell(row=1, column=2, value="Answer")
        self.sheet.cell(row=1, column=3, value="SQL_Query")
        self.sheet.cell(row=1, column=4, value="SQL_icon")
        self.sheet.cell(row=1, column=5, value="Expected_SQL_Query")
        self.sheet.cell(row=1, column=6, value="SQL_Query_Match")
        self.sheet.cell(row=1, column=7, value="similarity_score")
        self.sheet.cell(row=1, column=8, value="chart_icon")
        self.sheet.cell(row=1, column=9, value="actual_chart")
        self.sheet.cell(row=1, column=10, value="expected_chart_name")
        self.sheet.cell(row=1, column=11, value="chat_match")
        self.sheet.cell(row=1, column=12, value="Share_icon")
        self.sheet.cell(row=1, column=13, value="Save_icon")
        self.sheet.cell(row=1, column=14, value="View_fullscreen_icon")
        self.sheet.cell(row=1, column=15, value="data_view_icon")
        self.sheet.cell(row=1, column=16, value="Download_icon")
        self.sheet.cell(row=1, column=17, value="STATUS")


    def write_row(self, question, answer, sql_query, sql_icon,expected_sql_query, sql_query_match, similarity_score, chart_icon, actual_chart, expected_chart_name,chat_match, share_icon,save_icon,view_fullscreen_icon,data_view_icon, download_icon,status):
        self.row += 1
        self.sheet.cell(row=self.row, column=1, value=question)
        self.sheet.cell(row=self.row, column=2, value=answer)
        self.sheet.cell(row=self.row, column=3, value=sql_query)
        self.sheet.cell(row=self.row, column=4, value=sql_icon)
        self.sheet.cell(row=self.row, column=5, value=expected_sql_query)
        self.sheet.cell(row=self.row, column=6, value=sql_query_match)
        self.sheet.cell(row=self.row, column=7, value=similarity_score)
        self.sheet.cell(row=self.row, column=8, value=chart_icon)
        self.sheet.cell(row=self.row, column=9, value=actual_chart)
        self.sheet.cell(row=self.row, column=10, value=expected_chart_name)
        self.sheet.cell(row=self.row, column=11, value=chat_match)
        self.sheet.cell(row=self.row, column=12, value=share_icon)
        self.sheet.cell(row=self.row, column=13, value=save_icon)
        self.sheet.cell(row=self.row, column=14, value=view_fullscreen_icon)
        self.sheet.cell(row=self.row, column=15, value=data_view_icon)
        self.sheet.cell(row=self.row, column=16, value=download_icon)
        self.sheet.cell(row=self.row, column=17, value=status)

    def save(self):
        self.wb.save(self.file_path)
