from openpyxl import load_workbook

class ExcelReader:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.sheet = self.wb[sheet_name]

    def get_questions(self):
        questions = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):  # skip header
            if row[0]:
                questions.append(row[0]) # Column A
        return questions

    def get_sql_queries(self):
        sql_queries = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[1]:
                sql_queries.append(row[1])  # Column B
        return sql_queries

    def get_questions_with_expected_sql(self):
        data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            question = row[0]  # Column A
            expected_sql = row[1]  # Column B
            chart = row[3]  # Column D

            if question:
                data.append({
                    "question": question,
                    "expected_sql": expected_sql,
                    "chart": chart

                })
        return data